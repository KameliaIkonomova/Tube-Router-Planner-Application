import openpyxl
import collections
from collections import namedtuple, deque
from tkinter import *
import tkinter as tk
import tkinter.ttk as ttk

g = []
tube_line = []
tube_lines = []
show_course = []


def StartUp():
    master = tk.Tk()
    master.title("London Underground Route System Planner")
    master.geometry("325x95")
    master.configure(background="sky blue")
    # setting up the main menu GUI

    tk.Label(master, text="Where you coming from? ").grid(row=0)
    tk.Label(master, text="Where are you going to? ").grid(row=1)
    tk.Label(master, text="What time are you traveling at? ").grid(row=2)
    # asking the user questions

    e1 = tk.Entry(master)
    e2 = tk.Entry(master)
    e3 = ttk.Combobox(master)
    e3['values'] = (
    "5am", "6am", "7am", "8am", "9am", "10am", "11am", "12am", "1pm", "2pm", "3pm", "4pm", "5pm", "6pm", "7pm", "8pm",
    "9pm", "10pm", "11pm", "12pm")
    # getting the input from the user

    e1.grid(row=0, column=1)
    e2.grid(row=1, column=1)
    e3.grid(row=2, column=1)
    # formatting for the questions and input slots

    tk.Button(master, text='Okay', command=lambda: table(e1.get(), e2.get(), e3.get(), master)).grid(row=3, column=1,
                                                                                                     sticky=tk.W,
                                                                                                     pady=4)
    # button that takes you to the travel table

    tk.Button(master, text='Quit', command=lambda: quit()).grid(row=3, column=0, sticky=tk.W, pady=4)
    # button that lets you quit from the application

    tk.mainloop()


def search_Button(StartJourney, endJourney, time):
    StartJourney = StartJourney.title()
    endJourney = endJourney.title()
    # Puts a capital letter at the start of every user inputted word

    g = []
    tube_line = []
    tube_lines = []
    show_course = []
    # clearing the global variables

    if time == "5am" or time == "6am" or time == "7am" or time == "8am" or time == "5pm" or time == "6pm":
        ExcelSheet = 'London Underground data.xlsx'
    else:
        ExcelSheet = 'London Underground data-Bakerloo.xlsx'
    # this if statement takes the user inputted time and returns the correct excel sheet

    book = openpyxl.load_workbook(ExcelSheet)
    sheet = book.active

    # opening the excel sheet as a workbook

    class Node(object):
        # create a singly linked item
        def __init__(self, data=None, next=None, prev=None):
            self.next = next
            self.prev = prev
            self.data = data

    class DoublyLinkedList(object):
        # create an empty doubly linked list
        def __init__(self):
            self.head = None
            self.tail = None
            self.size = 0

        def append_item(self, data):
            # Append a item to the empty linked list
            item = Node(data, None, None)
            if self.head is None:
                self.head = item
                self.tail = self.head
            else:
                item.prev = self.tail
                self.tail.next = item
                self.tail = item

            self.size += 1

        def iter(self):
            # Iterate through the list
            current = self.head
            while current:
                item_val = current.data
                current = current.next
                yield item_val

        def reverse_iter(self):
            # Iterate backwards through the list
            current = self.tail
            while current:
                val = current.data
                current = current.prev
                yield val

        def print_foward(self):
            # Print items in list from first node inserted to the last
            for node in self.iter():
                print(node)

    tubestops = []
    tuple_list = []
    items = DoublyLinkedList()

    for row in sheet.iter_rows(min_row=1, min_col=2, max_row=770, max_col=4):

        tubestops_row = []
        tuple_row = ()

        for cell in row:
            tubestops_row.append(cell.value)
            tuple_row = tuple_row + (cell.value,)

        if tuple_row[2] is not None:
            tubestops.append(tubestops_row)
            tuple_list.append(tuple_row)
            items.append_item(tubestops_row)

            reversed_tubestops_row = tubestops_row[:]
            element0 = reversed_tubestops_row[0]
            reversed_tubestops_row[0] = reversed_tubestops_row[1]
            reversed_tubestops_row[1] = element0

            tubestops.append(reversed_tubestops_row)
            tuple_list.append(tuple(reversed_tubestops_row))
            items.append_item(reversed_tubestops_row)

    # using infinity as a default distance to items
    inf = float('inf')
    Edge = namedtuple('Edge', ['start', 'end', 'cost'])

    class Graph(object):
        def __init__(self, edges):
            self.edges = [Edge(*edge) for edge in edges]
            self.vertices = {e.start for e in self.edges} | {e.end for e in self.edges}

        def dijkstra(self, source, dest):
            assert source in self.vertices
            distances = {vertex: inf for vertex in self.vertices}
            previous = {vertex: None for vertex in self.vertices}
            distances[source] = 0
            p = self.vertices.copy()
            neighbours = {vertex: set() for vertex in self.vertices}
            for start, end, cost in self.edges:
                neighbours[start].add((end, cost))

            while p:
                cur_vert = min(p, key=lambda vertex: distances[vertex])
                p.remove(cur_vert)
                if distances[cur_vert] == inf or cur_vert == dest:
                    break
                for n, cost in neighbours[cur_vert]:
                    alt = distances[cur_vert] + cost
                    if alt < distances[n]:
                        distances[n] = alt
                        previous[n] = cur_vert
            road, cur_vert = deque(), dest
            while previous[cur_vert]:
                road.appendleft(cur_vert)
                cur_vert = previous[cur_vert]
            road.appendleft(cur_vert)
            return road

    graph = Graph(items.iter())
    lines = []
    var = []

    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=770, max_col=4):

        tubestops_row = []

        for cell in row:
            tubestops_row.append(cell.value)

        if tubestops_row[2] is not None:
            lines.append(tubestops_row)
    try:
        var_2 = collections.deque(graph.dijkstra(StartJourney, endJourney))
        if var_2 == None:
            pass

        while True:
            try:
                g.append(var_2.popleft())
            except IndexError:
                break

        col = collections.deque(lines)

        while True:
            try:
                var.append(col.popleft())
            except IndexError:
                break

        for i in range(len(g)):
            prev_tubestop = ""
            for row in var:
                if i < len(g) - 1:
                    if g[i] == row[1] and g[i + 1] == row[2] and row[1] is not prev_tubestop:
                        tube_line = [row[1], row[0], row[3]]
                        tube_lines.append(tube_line)
                        prev_tubestop = row[1]
                        LastStop = [row[2], row[0], 0]

        tube_lines.append(LastStop)

        for i in tube_lines:
            show_course.append(i[0:4])

        g.reverse()

        return show_course



    except:
        print("Incorrect Information", """Either:
    1. The entered tubestop is not available or incorrect
    2. Unnecessary space at the beginning or the end of the input""")


def table(start, end, time, master):
    master.destroy()
    # closing the main menu window

    route = search_Button(start, end, time)

    if route == None:
        root = Tk()
        ErrorMessage = tk.Label(root, text="Could not find start or destination, could be due to a spelling mistake or it not existing")
        ErrorMessage.grid(column=0, row=1)

        backButton = tk.Button(root, text="Back to Menu", command=lambda: restart(root))
        backButton.grid(column=0, row=2)
    else:
        totaltime = 0

        class Table:

            def __init__(self, root):
                for i in range(len(route) + 1):
                    for j in range(4):
                        self.e = Entry(root, width=25, fg="black", font=('Arial', 16, 'bold'))
                        self.e.grid(row=i, column=j)
                        self.e.insert(END, lst[i][j])

        # creating the information table

        lst = [("Station", "Line", "Travel time to next station", "Total travel time")]

        for time in route:
            time.append(totaltime)
            lst.append(time)
            totaltime += time[2] + 1
        # calculating the total journey time

        prevRoute = ""
        tubelineList = []
        for thisRoute in route:
            if (thisRoute[1] != prevRoute):
                tubelineList.append([thisRoute[0], thisRoute[1]])
            prevRoute = thisRoute[1]
        tubelineList.append([thisRoute[0], thisRoute[1]])
        # calculating the changes needed to make the journey

        root = Tk()
        # creating the root

        JourneySummary = tk.Label(root, text="Journey Summary")
        JourneySummary.grid(column=0, row=len(route) + 1)
        # Writing the title Journey Summary

        count = 1
        firstTime = True
        lineText = ''
        for change in tubelineList:

            if firstTime:
                lineText = str(change[1]) + " : " + str(change[0]) + " to "
                firstTime = False
            else:
                lineText = lineText + str(change[0])
                line = tk.Label(root, text=lineText)
                line.grid(column=0, row=len(route) + count)
                lineText = str(change[1]) + " : " + str(change[0]) + " to "
                count += 1
        # Writes the line changes you need to make

        JourneyTimeText = "Total journey time: ", str(totaltime), "mins"
        TotalJourneyTime = tk.Label(root, text=" ".join(JourneyTimeText))
        TotalJourneyTime.grid(column=0, row=len(route) + len(tubelineList) + 1)
        # Write the total journey time and formats it

        backButton = tk.Button(root, text="Back", command=lambda: restart(root))
        backButton.grid(column=0, row=len(route) + len(tubelineList) + 2)
        # Back button to get back to the main menu

        Table(root)
        root.mainloop()
        # running the root


def restart(root):
    root.destroy()
    StartUp()
    # function to destroy the journey table and open back up the main menu

StartUp()
